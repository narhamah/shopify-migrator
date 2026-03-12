#!/usr/bin/env python3
"""Comprehensive visual audit of the TARA Saudi Shopify store using Playwright.

Crawls the English and Arabic storefronts, checking for:
  - Untranslated text (English on Arabic pages, Spanish anywhere)
  - Broken images (404, missing src, missing alt)
  - Broken links (404, 5xx responses)
  - RTL layout issues (wrong text-direction, alignment anomalies)
  - Console errors (JS exceptions logged in browser)
  - Missing SEO meta tags (title, description, og:image)
  - Empty sections (visible containers with no content)
  - Accessibility issues (missing aria-labels, contrast hints)
  - Mobile responsiveness (viewport overflow, horizontal scroll)

Generates a detailed Excel report (.xlsx) with separate sheets per category.

Prerequisites:
    pip install playwright openpyxl python-dotenv
    playwright install chromium

Usage:
    python visual_audit_report.py --output data/audit.xlsx                     # Full audit (all pages, all checks)
    python visual_audit_report.py --output data/audit.xlsx --max-pages 20      # Quick scan (20 pages per locale)
    python visual_audit_report.py --output C:/Users/me/Desktop/audit.xlsx      # Custom output path
    python visual_audit_report.py --output data/audit.xlsx --no-screenshots    # Skip screenshots
    python visual_audit_report.py --output data/audit.xlsx --no-check-links    # Skip link checks (faster)
    python visual_audit_report.py --output data/audit.xlsx --english-only      # Skip Arabic checks
    python visual_audit_report.py --output data/audit.xlsx --arabic-only       # Skip English checks
"""

import argparse
import json
import os
import re
import sys
import time
from collections import defaultdict
from datetime import datetime
from urllib.parse import urlparse, urljoin

from dotenv import load_dotenv

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "src"))
from tara_migrate.core.language import (
    is_arabic_visible_text, count_chars, ARABIC_REGEX,
    find_untranslated_range_names,
)

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

TARA_OK_PATTERNS = [
    r"^(TARA|Tara|tara)(\s|$)",              # Brand name
    r"^(Kansa Wand|Gua Sha|Scalp Massager)",  # Product names (keep English)
    r"^ABG10\+®", r"^Capixyl™", r"^Procapil®", r"^Silverfree™",  # Trademarked ingredients
    r"^SAR\s", r"^SR\s",                       # Currency
    r"^\+966",                                 # Saudi phone
    r"^(Instagram|Facebook|TikTok|YouTube|X|WhatsApp)$",  # Social media names
    r"^\d+%?\s?(off|OFF)$",                    # Discount labels
    r"^(INCI|pH)\b",                           # Scientific abbreviations
    r"^[A-Z0-9]{2,5}$",                        # Short codes (SKUs, size codes)
    r"^©\s?\d{4}",                             # Copyright
    r"^(visa|mastercard|mada|apple pay|tabby)$",  # Payment method names (case-insensitive in check)
]

# Known-OK English phrases on Arabic pages (theme-level, hard to translate)
KNOWN_OK_ENGLISH = {
    "powered by shopify", "skip to content", "no reviews yet",
    "free shipping", "add to cart", "out of stock", "sold out",
}

# JS to extract all visible text with element metadata
EXTRACT_TEXT_JS = """() => {
    const results = [];
    const seen = new Set();
    const selectors = [
        'h1','h2','h3','h4','h5','h6','p','span','a','button','label',
        'li','td','th','figcaption','blockquote','cite','em','strong','b','i',
        'div:not(:has(*:not(br):not(wbr):not(img)))',
        '[class*="title"],[class*="heading"],[class*="label"],[class*="badge"]',
        '[class*="tab"],[class*="accordion"],[class*="btn"],[class*="price"]',
        '[class*="nav"],[class*="menu"],[class*="footer"],[class*="header"]',
        '[class*="description"],[class*="caption"],[class*="subtitle"]',
        '[data-testid],[role="heading"],[role="button"],[role="link"]',
    ].join(',');
    const elements = document.querySelectorAll(selectors);
    for (const el of elements) {
        const style = window.getComputedStyle(el);
        if (style.display==='none'||style.visibility==='hidden'||style.opacity==='0') continue;
        if (el.offsetWidth===0 && el.offsetHeight===0) continue;
        let text = '';
        for (const node of el.childNodes) {
            if (node.nodeType === Node.TEXT_NODE) text += node.textContent;
        }
        text = text.trim();
        if (!text && el.children.length === 0) text = (el.textContent||'').trim();
        if (!text || text.length < 2 || seen.has(text)) continue;
        seen.add(text);
        const rect = el.getBoundingClientRect();
        if (rect.width === 0 || rect.height === 0) continue;
        const tag = el.tagName.toLowerCase();
        results.push({
            text: text.substring(0, 500),
            tag,
            id: el.id || '',
            classes: el.className ? el.className.toString().substring(0, 150) : '',
            x: Math.round(rect.x),
            y: Math.round(rect.y),
            width: Math.round(rect.width),
            height: Math.round(rect.height),
            inViewport: rect.y < window.innerHeight * 5,
            dir: style.direction,
            textAlign: style.textAlign,
        });
    }
    return results;
}"""

# JS to extract image info
EXTRACT_IMAGES_JS = """() => {
    const imgs = [];
    document.querySelectorAll('img').forEach(img => {
        const rect = img.getBoundingClientRect();
        imgs.push({
            src: img.src || img.dataset.src || '',
            alt: img.alt || '',
            loading: img.loading || '',
            width: img.naturalWidth,
            height: img.naturalHeight,
            displayed: rect.width > 0 && rect.height > 0,
            classes: img.className ? img.className.toString().substring(0, 100) : '',
            parentTag: img.parentElement ? img.parentElement.tagName.toLowerCase() : '',
        });
    });
    // Also check background images
    document.querySelectorAll('[style*="background-image"]').forEach(el => {
        const bg = window.getComputedStyle(el).backgroundImage;
        const match = bg.match(/url\\(["']?([^"')]+)["']?\\)/);
        if (match) {
            const rect = el.getBoundingClientRect();
            imgs.push({
                src: match[1],
                alt: '(background-image)',
                loading: '',
                width: 0,
                height: 0,
                displayed: rect.width > 0 && rect.height > 0,
                classes: el.className ? el.className.toString().substring(0, 100) : '',
                parentTag: 'bg:' + el.tagName.toLowerCase(),
            });
        }
    });
    return imgs;
}"""

# JS to extract all links
EXTRACT_LINKS_JS = """() => {
    const links = [];
    const seen = new Set();
    document.querySelectorAll('a[href]').forEach(a => {
        const href = a.href;
        if (!href || seen.has(href)) return;
        seen.add(href);
        const rect = a.getBoundingClientRect();
        links.push({
            href,
            text: (a.textContent || '').trim().substring(0, 100),
            visible: rect.width > 0 && rect.height > 0,
            classes: a.className ? a.className.toString().substring(0, 100) : '',
            rel: a.rel || '',
        });
    });
    return links;
}"""

# JS to extract SEO meta tags
EXTRACT_SEO_JS = """() => {
    const meta = {};
    meta.title = document.title || '';
    const desc = document.querySelector('meta[name="description"]');
    meta.description = desc ? desc.content : '';
    const ogTitle = document.querySelector('meta[property="og:title"]');
    meta.og_title = ogTitle ? ogTitle.content : '';
    const ogDesc = document.querySelector('meta[property="og:description"]');
    meta.og_description = ogDesc ? ogDesc.content : '';
    const ogImage = document.querySelector('meta[property="og:image"]');
    meta.og_image = ogImage ? ogImage.content : '';
    const canonical = document.querySelector('link[rel="canonical"]');
    meta.canonical = canonical ? canonical.href : '';
    const hreflang = [];
    document.querySelectorAll('link[rel="alternate"][hreflang]').forEach(l => {
        hreflang.push({lang: l.hreflang, href: l.href});
    });
    meta.hreflang = hreflang;
    // Check for lang/dir attributes
    meta.html_lang = document.documentElement.lang || '';
    meta.html_dir = document.documentElement.dir || '';
    const h1s = document.querySelectorAll('h1');
    meta.h1_count = h1s.length;
    meta.h1_text = h1s.length > 0 ? h1s[0].textContent.trim().substring(0, 200) : '';
    return meta;
}"""

# JS to check for empty sections
EXTRACT_EMPTY_SECTIONS_JS = """() => {
    const results = [];
    const sectionSelectors = [
        'section', '[class*="section"]', '[class*="block"]',
        '[data-section-id]', '[data-section-type]',
        '.shopify-section',
    ];
    document.querySelectorAll(sectionSelectors.join(',')).forEach(el => {
        const rect = el.getBoundingClientRect();
        if (rect.height < 10 && rect.width > 100) {
            results.push({
                tag: el.tagName.toLowerCase(),
                id: el.id || '',
                classes: el.className ? el.className.toString().substring(0, 150) : '',
                height: Math.round(rect.height),
                width: Math.round(rect.width),
                sectionId: el.dataset.sectionId || '',
                sectionType: el.dataset.sectionType || '',
            });
        }
        // Also check for sections that are tall but completely empty of text
        const text = (el.textContent || '').trim();
        const imgs = el.querySelectorAll('img');
        if (rect.height > 50 && rect.width > 100 && !text && imgs.length === 0) {
            results.push({
                tag: el.tagName.toLowerCase(),
                id: el.id || '',
                classes: el.className ? el.className.toString().substring(0, 150) : '',
                height: Math.round(rect.height),
                width: Math.round(rect.width),
                sectionId: el.dataset.sectionId || '',
                sectionType: el.dataset.sectionType || '',
                reason: 'no text or images',
            });
        }
    });
    return results;
}"""

# JS to check RTL layout issues
CHECK_RTL_JS = """() => {
    const issues = [];
    // Check if page has RTL direction set
    const htmlDir = document.documentElement.dir;
    const bodyDir = document.body.dir || window.getComputedStyle(document.body).direction;
    if (htmlDir !== 'rtl' && bodyDir !== 'rtl') {
        issues.push({type: 'missing_rtl', detail: 'Page does not have dir="rtl"'});
    }
    // Check for elements with explicit LTR that might be wrong
    const ltrElements = document.querySelectorAll('[dir="ltr"]');
    for (const el of ltrElements) {
        const text = (el.textContent || '').trim();
        // Check if it has Arabic text but is forced LTR
        const arabicChars = (text.match(/[\\u0600-\\u06FF]/g) || []).length;
        const totalChars = text.replace(/[^a-zA-Z\\u0600-\\u06FF]/g, '').length;
        if (totalChars > 5 && arabicChars / totalChars > 0.5) {
            issues.push({
                type: 'arabic_in_ltr',
                detail: text.substring(0, 100),
                tag: el.tagName.toLowerCase(),
                classes: el.className ? el.className.toString().substring(0, 100) : '',
            });
        }
    }
    // Check for text-align:left on elements with Arabic text
    const textElements = document.querySelectorAll('p, h1, h2, h3, h4, h5, h6, span, div, li');
    for (const el of textElements) {
        const style = window.getComputedStyle(el);
        if (style.textAlign === 'left' && style.direction === 'rtl') {
            const text = (el.textContent || '').trim();
            const arabicChars = (text.match(/[\\u0600-\\u06FF]/g) || []).length;
            if (arabicChars > 10) {
                issues.push({
                    type: 'left_align_rtl',
                    detail: text.substring(0, 100),
                    tag: el.tagName.toLowerCase(),
                    classes: el.className ? el.className.toString().substring(0, 100) : '',
                });
                if (issues.length > 50) break;
            }
        }
    }
    // Check for horizontal overflow
    if (document.body.scrollWidth > window.innerWidth + 5) {
        issues.push({
            type: 'horizontal_overflow',
            detail: `Body scrollWidth (${document.body.scrollWidth}) > viewport (${window.innerWidth})`,
        });
    }
    return issues;
}"""

# JS to check accessibility
CHECK_A11Y_JS = """() => {
    const issues = [];
    // Buttons without accessible name
    document.querySelectorAll('button').forEach(btn => {
        const text = (btn.textContent || '').trim();
        const ariaLabel = btn.getAttribute('aria-label') || '';
        const title = btn.title || '';
        if (!text && !ariaLabel && !title) {
            issues.push({
                type: 'button_no_label',
                classes: btn.className ? btn.className.toString().substring(0, 100) : '',
                html: btn.outerHTML.substring(0, 200),
            });
        }
    });
    // Links without text
    document.querySelectorAll('a').forEach(a => {
        const text = (a.textContent || '').trim();
        const ariaLabel = a.getAttribute('aria-label') || '';
        const title = a.title || '';
        const img = a.querySelector('img[alt]');
        if (!text && !ariaLabel && !title && !img) {
            const rect = a.getBoundingClientRect();
            if (rect.width > 0 && rect.height > 0) {
                issues.push({
                    type: 'link_no_text',
                    href: a.href || '',
                    classes: a.className ? a.className.toString().substring(0, 100) : '',
                });
            }
        }
    });
    // Form inputs without labels
    document.querySelectorAll('input:not([type="hidden"]), select, textarea').forEach(inp => {
        const id = inp.id;
        const ariaLabel = inp.getAttribute('aria-label') || '';
        const placeholder = inp.placeholder || '';
        const label = id ? document.querySelector(`label[for="${id}"]`) : null;
        if (!label && !ariaLabel && !placeholder) {
            issues.push({
                type: 'input_no_label',
                inputType: inp.type || inp.tagName.toLowerCase(),
                name: inp.name || '',
                classes: inp.className ? inp.className.toString().substring(0, 100) : '',
            });
        }
    });
    return issues;
}"""

# JS to expand accordions and tabs
EXPAND_INTERACTIVE_JS = """() => {
    // Accordions
    document.querySelectorAll(
        '[data-accordion], .accordion__trigger, details summary, ' +
        '[class*="accordion"] button, [class*="collapse"] button, ' +
        '[aria-expanded="false"]'
    ).forEach(el => {
        try { el.click(); } catch(e) {}
    });
    // Tabs — click all tab buttons
    document.querySelectorAll(
        '[role="tab"], [class*="tab"] button, [class*="tab-trigger"]'
    ).forEach(el => {
        try { el.click(); } catch(e) {}
    });
}"""

# JS to check mobile viewport
CHECK_MOBILE_JS = """() => {
    const issues = [];
    // Check for elements wider than viewport
    const viewportWidth = window.innerWidth;
    const allElements = document.querySelectorAll('*');
    const overflowing = new Set();
    for (const el of allElements) {
        const rect = el.getBoundingClientRect();
        if (rect.right > viewportWidth + 5 && rect.width > 50) {
            const tag = el.tagName.toLowerCase();
            const cls = el.className ? el.className.toString().substring(0, 80) : '';
            const key = tag + '.' + cls;
            if (!overflowing.has(key)) {
                overflowing.add(key);
                issues.push({
                    tag,
                    classes: cls,
                    width: Math.round(rect.width),
                    overflow: Math.round(rect.right - viewportWidth),
                });
            }
            if (issues.length > 20) break;
        }
    }
    return issues;
}"""


# ─────────────────────────────────────────────────────────────────────────────
# Issue collectors
# ─────────────────────────────────────────────────────────────────────────────

def _is_spanish(text):
    """Check if text contains Spanish morphology indicators."""
    indicators = re.findall(
        r"(?:ción|iones|amente|amiento|ular|ficante|para\s|con\s|del\s|los\s|las\s|una?\s)",
        text, re.IGNORECASE,
    )
    return len(indicators) >= 2


def check_untranslated_text(texts, locale, url):
    """Find text that should be translated but isn't."""
    issues = []
    for t in texts:
        text = t["text"]
        if not t.get("inViewport"):
            continue

        if locale == "ar":
            # On Arabic pages: flag English/Spanish text
            lower = text.lower().strip()
            if lower in KNOWN_OK_ENGLISH:
                continue
            if not is_arabic_visible_text(text, ok_patterns=TARA_OK_PATTERNS):
                severity = "high" if len(text) > 20 else "medium"
                # Extra severity if it's a heading
                if t["tag"] in ("h1", "h2", "h3"):
                    severity = "critical"
                issues.append({
                    "category": "Untranslated Text",
                    "severity": severity,
                    "url": url,
                    "page_type": _classify_page(url),
                    "element": f"<{t['tag']}>",
                    "css_classes": t.get("classes", ""),
                    "text": text[:300],
                    "position": f"({t['x']}, {t['y']})",
                    "detail": "English/Latin text on Arabic page",
                })
            else:
                # Text passes Arabic ratio check, but may contain embedded
                # English range/collection names that should be translated
                range_hits = find_untranslated_range_names(text)
                for en_name, ar_name in range_hits:
                    issues.append({
                        "category": "Untranslated Text",
                        "severity": "high",
                        "url": url,
                        "page_type": _classify_page(url),
                        "element": f"<{t['tag']}>",
                        "css_classes": t.get("classes", ""),
                        "text": text[:300],
                        "position": f"({t['x']}, {t['y']})",
                        "detail": f"English range name '{en_name}' in Arabic text — should be '{ar_name}'",
                    })
        else:
            # On English pages: flag Spanish text
            if _is_spanish(text):
                issues.append({
                    "category": "Untranslated Text",
                    "severity": "high",
                    "url": url,
                    "page_type": _classify_page(url),
                    "element": f"<{t['tag']}>",
                    "css_classes": t.get("classes", ""),
                    "text": text[:300],
                    "position": f"({t['x']}, {t['y']})",
                    "detail": "Spanish text on English page",
                })
    return issues


def check_images(images, url):
    """Check for image issues."""
    issues = []
    for img in images:
        src = img.get("src", "")

        # Missing src
        if not src or src == "about:blank":
            if img.get("displayed"):
                issues.append({
                    "category": "Broken Image",
                    "severity": "high",
                    "url": url,
                    "page_type": _classify_page(url),
                    "element": "<img>",
                    "css_classes": img.get("classes", ""),
                    "text": f"alt='{img.get('alt', '')}'",
                    "position": "",
                    "detail": "Image displayed but has no src",
                })
            continue

        # Broken image (loaded but 0 dimensions = failed load)
        if img.get("displayed") and img.get("width", 0) == 0 and img.get("height", 0) == 0:
            if not src.startswith("data:") and "background-image" not in img.get("parentTag", ""):
                issues.append({
                    "category": "Broken Image",
                    "severity": "high",
                    "url": url,
                    "page_type": _classify_page(url),
                    "element": "<img>",
                    "css_classes": img.get("classes", ""),
                    "text": src[:200],
                    "position": "",
                    "detail": "Image failed to load (0x0 dimensions)",
                })

        # Missing alt text
        alt = img.get("alt", "").strip()
        if not alt and img.get("displayed") and not src.startswith("data:"):
            parent = img.get("parentTag", "")
            # Skip decorative images (in links, bg)
            if parent not in ("a", "button") and "bg:" not in parent:
                issues.append({
                    "category": "Missing Alt Text",
                    "severity": "low",
                    "url": url,
                    "page_type": _classify_page(url),
                    "element": "<img>",
                    "css_classes": img.get("classes", ""),
                    "text": src[:200],
                    "position": "",
                    "detail": "Image has no alt text (accessibility)",
                })
    return issues


def check_seo(meta, url, locale):
    """Check SEO meta tags."""
    issues = []
    page_type = _classify_page(url)

    if not meta.get("title"):
        issues.append({
            "category": "SEO",
            "severity": "critical",
            "url": url,
            "page_type": page_type,
            "element": "<title>",
            "css_classes": "",
            "text": "(empty)",
            "position": "",
            "detail": "Page has no title tag",
        })

    if not meta.get("description"):
        issues.append({
            "category": "SEO",
            "severity": "high",
            "url": url,
            "page_type": page_type,
            "element": "meta[description]",
            "css_classes": "",
            "text": "(empty)",
            "position": "",
            "detail": "Missing meta description",
        })

    if not meta.get("og_image"):
        issues.append({
            "category": "SEO",
            "severity": "medium",
            "url": url,
            "page_type": page_type,
            "element": "meta[og:image]",
            "css_classes": "",
            "text": "(empty)",
            "position": "",
            "detail": "Missing Open Graph image",
        })

    h1_count = meta.get("h1_count", 0)
    if h1_count == 0:
        issues.append({
            "category": "SEO",
            "severity": "high",
            "url": url,
            "page_type": page_type,
            "element": "<h1>",
            "css_classes": "",
            "text": "(none)",
            "position": "",
            "detail": "Page has no H1 tag",
        })
    elif h1_count > 1:
        issues.append({
            "category": "SEO",
            "severity": "low",
            "url": url,
            "page_type": page_type,
            "element": "<h1>",
            "css_classes": "",
            "text": meta.get("h1_text", "")[:100],
            "position": "",
            "detail": f"Page has {h1_count} H1 tags (should be 1)",
        })

    # Check title/desc language matches locale
    if locale == "ar":
        title = meta.get("title", "")
        if title and len(title) > 5:
            arabic, latin = count_chars(title)
            total = arabic + latin
            if total > 5 and arabic / total < 0.3:
                issues.append({
                    "category": "SEO",
                    "severity": "high",
                    "url": url,
                    "page_type": page_type,
                    "element": "<title>",
                    "css_classes": "",
                    "text": title[:200],
                    "position": "",
                    "detail": "Title tag appears to be in English, not Arabic",
                })

    # Check html lang attribute
    expected_lang = "ar" if locale == "ar" else "en"
    html_lang = meta.get("html_lang", "")
    if html_lang and not html_lang.startswith(expected_lang):
        issues.append({
            "category": "SEO",
            "severity": "medium",
            "url": url,
            "page_type": page_type,
            "element": "<html lang>",
            "css_classes": "",
            "text": f"lang='{html_lang}'",
            "position": "",
            "detail": f"Expected lang='{expected_lang}*', got '{html_lang}'",
        })

    return issues


def check_empty_sections(sections, url):
    """Check for empty/collapsed sections."""
    issues = []
    for s in sections:
        issues.append({
            "category": "Empty Section",
            "severity": "medium",
            "url": url,
            "page_type": _classify_page(url),
            "element": f"<{s['tag']}>",
            "css_classes": s.get("classes", ""),
            "text": s.get("sectionType") or s.get("sectionId") or s.get("id") or "(unnamed)",
            "position": f"h={s['height']}px, w={s['width']}px",
            "detail": s.get("reason", f"Section collapsed to {s['height']}px height"),
        })
    return issues


def check_rtl_issues(rtl_data, url):
    """Check RTL layout problems."""
    issues = []
    for item in rtl_data:
        severity_map = {
            "missing_rtl": "critical",
            "arabic_in_ltr": "high",
            "left_align_rtl": "medium",
            "horizontal_overflow": "high",
        }
        issues.append({
            "category": "RTL Layout",
            "severity": severity_map.get(item["type"], "medium"),
            "url": url,
            "page_type": _classify_page(url),
            "element": f"<{item.get('tag', 'html')}>",
            "css_classes": item.get("classes", ""),
            "text": item.get("detail", "")[:300],
            "position": "",
            "detail": item["type"].replace("_", " ").title(),
        })
    return issues


def check_accessibility(a11y_data, url):
    """Check accessibility issues."""
    issues = []
    for item in a11y_data:
        type_detail = {
            "button_no_label": ("Button has no accessible name", "medium"),
            "link_no_text": ("Link has no text or aria-label", "medium"),
            "input_no_label": ("Form input has no label", "medium"),
        }
        detail, severity = type_detail.get(item["type"], (item["type"], "low"))
        issues.append({
            "category": "Accessibility",
            "severity": severity,
            "url": url,
            "page_type": _classify_page(url),
            "element": item.get("tag", item["type"]),
            "css_classes": item.get("classes", ""),
            "text": item.get("html", item.get("href", item.get("name", "")))[:200],
            "position": "",
            "detail": detail,
        })
    return issues


def check_console_errors(errors, url):
    """Convert captured console errors to issues."""
    issues = []
    for err in errors:
        issues.append({
            "category": "Console Error",
            "severity": "medium",
            "url": url,
            "page_type": _classify_page(url),
            "element": "console",
            "css_classes": "",
            "text": err[:300],
            "position": "",
            "detail": "JavaScript error in browser console",
        })
    return issues


def check_mobile_overflow(overflow_data, url):
    """Check mobile viewport overflow."""
    issues = []
    for item in overflow_data:
        issues.append({
            "category": "Mobile Overflow",
            "severity": "medium",
            "url": url,
            "page_type": _classify_page(url),
            "element": f"<{item['tag']}>",
            "css_classes": item.get("classes", ""),
            "text": f"width={item['width']}px, overflow={item['overflow']}px",
            "position": "",
            "detail": "Element overflows mobile viewport",
        })
    return issues


def _classify_page(url):
    """Classify a URL into a page type."""
    path = urlparse(url).path
    # Strip locale prefix
    path = re.sub(r"^/(ar|en)(/|$)", "/", path)
    if path in ("/", ""):
        return "Homepage"
    if "/products/" in path:
        return "Product"
    if "/collections/" in path:
        return "Collection"
    if "/pages/" in path:
        return "Page"
    if "/blogs/" in path and path.count("/") > 2:
        return "Article"
    if "/blogs/" in path:
        return "Blog"
    if "/cart" in path:
        return "Cart"
    if "/search" in path:
        return "Search"
    if "/account" in path:
        return "Account"
    return "Other"


# ─────────────────────────────────────────────────────────────────────────────
# Link checker (batch)
# ─────────────────────────────────────────────────────────────────────────────

def check_links_batch(page, all_links, domain):
    """Check a batch of links for broken responses. Returns issues list."""
    issues = []
    checked = set()
    for link_info in all_links:
        href = link_info["href"]
        source_url = link_info.get("source_url", "")

        if href in checked:
            continue
        checked.add(href)

        parsed = urlparse(href)
        # Only check same-domain links
        if parsed.netloc and parsed.netloc != domain:
            continue
        # Skip anchors, javascript, mailto, tel
        if href.startswith(("javascript:", "mailto:", "tel:", "#")):
            continue

        try:
            resp = page.request.get(href, timeout=10000)
            status = resp.status
            if status >= 400:
                severity = "critical" if status == 404 else "high"
                issues.append({
                    "category": "Broken Link",
                    "severity": severity,
                    "url": source_url,
                    "page_type": _classify_page(source_url),
                    "element": "<a>",
                    "css_classes": link_info.get("classes", ""),
                    "text": f"{link_info.get('text', '')} → {href}"[:300],
                    "position": "",
                    "detail": f"HTTP {status}",
                })
        except Exception:
            pass  # Timeout or network error — skip

    return issues


# ─────────────────────────────────────────────────────────────────────────────
# Crawler
# ─────────────────────────────────────────────────────────────────────────────

def audit_page(page, url, locale, console_errors, screenshots_dir=None):
    """Run all checks on a single page. Returns list of issue dicts."""
    issues = []
    path = urlparse(url).path

    try:
        response = page.goto(url, wait_until="networkidle", timeout=30000)
        if not response:
            return [{"category": "Page Error", "severity": "critical", "url": url,
                     "page_type": _classify_page(url), "element": "", "css_classes": "",
                     "text": "No response", "position": "", "detail": "Page returned no response"}]
        if response.status >= 400:
            return [{"category": "Page Error", "severity": "critical", "url": url,
                     "page_type": _classify_page(url), "element": "", "css_classes": "",
                     "text": f"HTTP {response.status}", "position": "",
                     "detail": f"Page returned HTTP {response.status}"}]

        time.sleep(1)

        # Expand accordions/tabs to reveal hidden content
        try:
            page.evaluate(EXPAND_INTERACTIVE_JS)
            time.sleep(0.5)
        except Exception:
            pass

        # Screenshot
        if screenshots_dir:
            fname = path.strip("/").replace("/", "_") or "home"
            if len(fname) > 100:
                fname = fname[:100]
            try:
                page.screenshot(
                    path=os.path.join(screenshots_dir, f"{fname}.png"),
                    full_page=True,
                )
            except Exception:
                pass

        # Run all checks
        texts = page.evaluate(EXTRACT_TEXT_JS)
        issues.extend(check_untranslated_text(texts, locale, url))

        images = page.evaluate(EXTRACT_IMAGES_JS)
        issues.extend(check_images(images, url))

        seo_meta = page.evaluate(EXTRACT_SEO_JS)
        issues.extend(check_seo(seo_meta, url, locale))

        empty = page.evaluate(EXTRACT_EMPTY_SECTIONS_JS)
        issues.extend(check_empty_sections(empty, url))

        if locale == "ar":
            rtl = page.evaluate(CHECK_RTL_JS)
            issues.extend(check_rtl_issues(rtl, url))

        a11y = page.evaluate(CHECK_A11Y_JS)
        issues.extend(check_accessibility(a11y, url))

        # Console errors captured by listener
        if console_errors:
            issues.extend(check_console_errors(list(console_errors), url))
            console_errors.clear()

        # Extract navigation links for crawling
        links = page.evaluate(EXTRACT_LINKS_JS)

    except Exception as e:
        return [{"category": "Page Error", "severity": "high", "url": url,
                 "page_type": _classify_page(url), "element": "", "css_classes": "",
                 "text": str(e)[:300], "position": "", "detail": "Error loading page"}]

    return issues, links


def crawl_and_audit(base_url, locale, page, max_pages=100,
                    screenshots_dir=None):
    """Crawl the site and audit every page. Returns (all_issues, visited_urls, all_links)."""
    visited = set()
    domain = urlparse(base_url).netloc
    locale_prefix = f"/{locale}" if locale != "en" else ""

    # Seed URLs
    home = base_url.rstrip("/")
    to_visit = [
        home,
        home + "/collections/all",
        home + "/collections",
        home + "/pages/ingredients",
        home + "/blogs/journal",
    ]

    all_issues = []
    all_links = []
    page_count = 0
    console_errors = []

    # Capture console errors
    page.on("console", lambda msg: (
        console_errors.append(msg.text)
        if msg.type == "error" and "favicon" not in msg.text.lower()
        else None
    ))

    while to_visit and page_count < max_pages:
        url = to_visit.pop(0)
        if url in visited:
            continue

        parsed = urlparse(url)
        if parsed.netloc != domain:
            continue
        # Enforce locale prefix for non-English
        if locale_prefix and locale_prefix not in parsed.path:
            continue

        visited.add(url)
        page_count += 1
        short_path = parsed.path

        label = f"{page_count}" if max_pages > 99999 else f"{page_count}/{max_pages}"
        print(f"  [{label}] {short_path}")

        result = audit_page(page, url, locale, console_errors, screenshots_dir)
        if isinstance(result, list):
            # Error case — no links returned
            all_issues.extend(result)
            continue

        page_issues, links = result

        if page_issues:
            # Count by category
            by_cat = defaultdict(int)
            for issue in page_issues:
                by_cat[issue["category"]] += 1
            summary = ", ".join(f"{c}: {n}" for c, n in sorted(by_cat.items()))
            print(f"           {summary}")

        all_issues.extend(page_issues)

        # Add discovered links to crawl queue
        for link in links:
            href = link["href"].split("#")[0].split("?")[0]
            if href not in visited and domain in href:
                if locale_prefix and locale_prefix in urlparse(href).path:
                    to_visit.append(href)
                    all_links.append({**link, "source_url": url})
                elif not locale_prefix:
                    to_visit.append(href)
                    all_links.append({**link, "source_url": url})

    return all_issues, visited, all_links


# ─────────────────────────────────────────────────────────────────────────────
# Mobile viewport audit
# ─────────────────────────────────────────────────────────────────────────────

def audit_mobile(base_url, locale, browser, max_pages=10):
    """Quick mobile audit on a subset of pages."""
    context = browser.new_context(
        viewport={"width": 375, "height": 812},  # iPhone size
        user_agent="Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15",
        locale="ar-SA" if locale == "ar" else "en-US",
    )
    page = context.new_page()
    issues = []

    home = base_url.rstrip("/")
    urls = [
        home,
        home + "/collections/all",
    ]
    # Discover a few product pages
    try:
        page.goto(home + "/collections/all", wait_until="networkidle", timeout=30000)
        time.sleep(1)
        product_links = page.evaluate("""() => {
            const links = [];
            document.querySelectorAll('a[href*="/products/"]').forEach(a => {
                const href = a.href.split('#')[0].split('?')[0];
                if (!links.includes(href)) links.push(href);
            });
            return links.slice(0, 5);
        }""")
        urls.extend(product_links)
    except Exception:
        pass

    for url in urls[:max_pages]:
        path = urlparse(url).path
        print(f"  [mobile] {path}")
        try:
            page.goto(url, wait_until="networkidle", timeout=30000)
            time.sleep(1)
            overflow = page.evaluate(CHECK_MOBILE_JS)
            issues.extend(check_mobile_overflow(overflow, url))
        except Exception as e:
            print(f"           Error: {e}")

    context.close()
    return issues


# ─────────────────────────────────────────────────────────────────────────────
# Excel report writer
# ─────────────────────────────────────────────────────────────────────────────

def write_excel_report(all_issues, output_path, visited_urls, audit_meta):
    """Write issues to a formatted Excel workbook with multiple sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Color scheme ──
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    CRITICAL_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    CRITICAL_FONT = Font(color="FFFFFF", bold=True)
    HIGH_FILL = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    HIGH_FONT = Font(color="FFFFFF")
    MEDIUM_FILL = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
    LOW_FILL = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    severity_styles = {
        "critical": (CRITICAL_FILL, CRITICAL_FONT),
        "high": (HIGH_FILL, HIGH_FONT),
        "medium": (MEDIUM_FILL, Font()),
        "low": (LOW_FILL, Font()),
    }

    COLUMNS = ["Category", "Severity", "Page Type", "URL", "Element",
               "CSS Classes", "Text / Content", "Position", "Detail"]

    def style_sheet(ws, rows):
        """Apply styling to a worksheet."""
        # Header row
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        # Data rows
        for row_idx, issue in enumerate(rows, 2):
            values = [
                issue["category"], issue["severity"], issue["page_type"],
                issue["url"], issue["element"], issue["css_classes"],
                issue["text"], issue["position"], issue["detail"],
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=col_idx in (7, 9))
                # Color severity column
                if col_idx == 2:
                    fill, font = severity_styles.get(issue["severity"], (LOW_FILL, Font()))
                    cell.fill = fill
                    cell.font = font
                    cell.alignment = Alignment(horizontal="center", vertical="top")

        # Auto-width columns (cap at 60)
        for col_idx in range(1, len(COLUMNS) + 1):
            max_len = len(COLUMNS[col_idx - 1])
            for row_idx in range(2, min(len(rows) + 2, 50)):
                cell_val = ws.cell(row=row_idx, column=col_idx).value or ""
                max_len = max(max_len, min(len(str(cell_val)), 60))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

        # Freeze header
        ws.freeze_panes = "A2"
        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"

    # ── 1. Summary sheet ──
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Title
    ws_summary.merge_cells("A1:D1")
    title_cell = ws_summary["A1"]
    title_cell.value = "TARA Saudi Store — Visual Audit Report"
    title_cell.font = Font(size=16, bold=True, color="1F4E79")

    ws_summary["A3"] = "Generated"
    ws_summary["B3"] = audit_meta.get("timestamp", "")
    ws_summary["A4"] = "Base URL"
    ws_summary["B4"] = audit_meta.get("base_url", "")
    ws_summary["A5"] = "Locale"
    ws_summary["B5"] = audit_meta.get("locale", "")
    ws_summary["A6"] = "Pages Crawled"
    ws_summary["B6"] = len(visited_urls)
    ws_summary["A7"] = "Total Issues"
    ws_summary["B7"] = len(all_issues)
    for r in range(3, 8):
        ws_summary.cell(row=r, column=1).font = Font(bold=True)

    # Breakdown by category
    ws_summary["A9"] = "Issues by Category"
    ws_summary["A9"].font = Font(size=13, bold=True, color="1F4E79")
    by_cat = defaultdict(lambda: {"critical": 0, "high": 0, "medium": 0, "low": 0, "total": 0})
    for issue in all_issues:
        cat = issue["category"]
        sev = issue["severity"]
        by_cat[cat][sev] += 1
        by_cat[cat]["total"] += 1

    cat_headers = ["Category", "Critical", "High", "Medium", "Low", "Total"]
    for col_idx, h in enumerate(cat_headers, 1):
        cell = ws_summary.cell(row=10, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for row_idx, (cat, counts) in enumerate(sorted(by_cat.items(), key=lambda x: -x[1]["total"]), 11):
        ws_summary.cell(row=row_idx, column=1, value=cat).border = THIN_BORDER
        ws_summary.cell(row=row_idx, column=2, value=counts["critical"]).border = THIN_BORDER
        ws_summary.cell(row=row_idx, column=3, value=counts["high"]).border = THIN_BORDER
        ws_summary.cell(row=row_idx, column=4, value=counts["medium"]).border = THIN_BORDER
        ws_summary.cell(row=row_idx, column=5, value=counts["low"]).border = THIN_BORDER
        ws_summary.cell(row=row_idx, column=6, value=counts["total"]).border = THIN_BORDER
        ws_summary.cell(row=row_idx, column=6).font = Font(bold=True)

    # Breakdown by page type
    row_offset = 11 + len(by_cat) + 2
    ws_summary.cell(row=row_offset, column=1, value="Issues by Page Type").font = Font(size=13, bold=True, color="1F4E79")
    by_page = defaultdict(int)
    for issue in all_issues:
        by_page[issue["page_type"]] += 1

    for col_idx, h in enumerate(["Page Type", "Issue Count"], 1):
        cell = ws_summary.cell(row=row_offset + 1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for row_idx, (pt, count) in enumerate(sorted(by_page.items(), key=lambda x: -x[1]), row_offset + 2):
        ws_summary.cell(row=row_idx, column=1, value=pt).border = THIN_BORDER
        ws_summary.cell(row=row_idx, column=2, value=count).border = THIN_BORDER

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 50

    # ── 2. All Issues sheet ──
    ws_all = wb.create_sheet("All Issues")
    # Sort: critical first, then high, medium, low
    severity_order = {"critical": 0, "high": 1, "medium": 2, "low": 3}
    sorted_issues = sorted(all_issues, key=lambda x: (severity_order.get(x["severity"], 9), x["category"]))
    style_sheet(ws_all, sorted_issues)

    # ── 3. Per-category sheets ──
    category_order = [
        "Untranslated Text", "Broken Link", "Broken Image", "Page Error",
        "SEO", "RTL Layout", "Empty Section", "Console Error",
        "Missing Alt Text", "Accessibility", "Mobile Overflow",
    ]
    for cat in category_order:
        cat_issues = [i for i in sorted_issues if i["category"] == cat]
        if not cat_issues:
            continue
        # Sanitize sheet name (max 31 chars, no special chars)
        sheet_name = cat[:31]
        ws = wb.create_sheet(sheet_name)
        style_sheet(ws, cat_issues)

    # ── 4. Pages Crawled sheet ──
    ws_pages = wb.create_sheet("Pages Crawled")
    ws_pages.cell(row=1, column=1, value="URL").font = HEADER_FONT
    ws_pages["A1"].fill = HEADER_FILL
    ws_pages.cell(row=1, column=2, value="Page Type").font = HEADER_FONT
    ws_pages["B1"].fill = HEADER_FILL
    ws_pages.cell(row=1, column=3, value="Issues").font = HEADER_FONT
    ws_pages["C1"].fill = HEADER_FILL

    issues_per_url = defaultdict(int)
    for issue in all_issues:
        issues_per_url[issue["url"]] += 1

    for row_idx, url in enumerate(sorted(visited_urls), 2):
        ws_pages.cell(row=row_idx, column=1, value=url)
        ws_pages.cell(row=row_idx, column=2, value=_classify_page(url))
        ws_pages.cell(row=row_idx, column=3, value=issues_per_url.get(url, 0))

    ws_pages.column_dimensions["A"].width = 80
    ws_pages.column_dimensions["B"].width = 15
    ws_pages.column_dimensions["C"].width = 10
    ws_pages.freeze_panes = "A2"

    # Save
    wb.save(output_path)
    print(f"\nExcel report saved to: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Comprehensive visual audit of TARA Saudi Shopify store → Excel report")
    parser.add_argument("--base-url", default="https://sa.taraformula.com",
                        help="Base URL of the store")
    parser.add_argument("--max-pages", type=int, default=0,
                        help="Max pages to crawl per locale (0 = unlimited, default: unlimited)")
    parser.add_argument("--output", required=True,
                        help="Output Excel file path (e.g. data/audit_report.xlsx)")
    parser.add_argument("--no-screenshots", action="store_true",
                        help="Skip saving full-page screenshots (screenshots ON by default)")
    parser.add_argument("--headed", action="store_true",
                        help="Run with visible browser")
    parser.add_argument("--english-only", action="store_true",
                        help="Only audit English pages")
    parser.add_argument("--arabic-only", action="store_true",
                        help="Only audit Arabic pages")
    parser.add_argument("--no-check-links", action="store_true",
                        help="Skip broken link checking (link checks ON by default)")
    parser.add_argument("--no-mobile", action="store_true",
                        help="Skip mobile viewport checks (mobile checks ON by default)")
    parser.add_argument("--json-out", default=None,
                        help="Also save raw JSON results")
    args = parser.parse_args()

    load_dotenv()

    # Ensure output directory exists
    output_dir = os.path.dirname(args.output)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    # 0 = unlimited — use a very high number
    if args.max_pages <= 0:
        args.max_pages = 999999

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("ERROR: pip install playwright && playwright install chromium")
        sys.exit(1)

    locales = []
    if not args.english_only:
        locales.append(("ar", "/ar"))
    if not args.arabic_only:
        locales.append(("en", ""))

    all_issues = []
    all_visited = set()

    print("=" * 70)
    print("  TARA SAUDI STORE — COMPREHENSIVE VISUAL AUDIT")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 70)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=not args.headed)

        for locale, prefix in locales:
            locale_label = locale.upper()
            browser_locale = "ar-SA" if locale == "ar" else "en-US"

            print(f"\n{'─' * 70}")
            print(f"  CRAWLING {locale_label} PAGES")
            print(f"{'─' * 70}")

            context = browser.new_context(
                viewport={"width": 1440, "height": 900},
                locale=browser_locale,
            )
            page = context.new_page()

            screenshots_dir = None
            if not args.no_screenshots:
                # Store screenshots next to the output file
                output_base = os.path.splitext(args.output)[0]
                screenshots_dir = f"{output_base}_screenshots_{locale}"
                os.makedirs(screenshots_dir, exist_ok=True)

            start_url = f"{args.base_url.rstrip('/')}{prefix}"
            issues, visited, links = crawl_and_audit(
                start_url, locale, page,
                max_pages=args.max_pages,
                screenshots_dir=screenshots_dir,
            )

            # Check broken links (on by default)
            if not args.no_check_links and links:
                print(f"\n  Checking {len(links)} links for broken responses...")
                domain = urlparse(args.base_url).netloc
                link_issues = check_links_batch(page, links, domain)
                issues.extend(link_issues)
                print(f"  Found {len(link_issues)} broken links")

            all_issues.extend(issues)
            all_visited.update(visited)

            context.close()

            # Print locale summary
            by_sev = defaultdict(int)
            for i in issues:
                by_sev[i["severity"]] += 1
            print(f"\n  {locale_label} Summary: {len(visited)} pages, {len(issues)} issues")
            print(f"    Critical: {by_sev['critical']}  High: {by_sev['high']}  "
                  f"Medium: {by_sev['medium']}  Low: {by_sev['low']}")

        # Mobile audit (on by default)
        if not args.no_mobile:
            print(f"\n{'─' * 70}")
            print("  MOBILE VIEWPORT AUDIT")
            print(f"{'─' * 70}")
            for locale, prefix in locales:
                start_url = f"{args.base_url.rstrip('/')}{prefix}"
                mobile_issues = audit_mobile(start_url, locale, browser, max_pages=20)
                all_issues.extend(mobile_issues)
                print(f"  {locale.upper()}: {len(mobile_issues)} mobile overflow issues")

        browser.close()

    # ── Generate report ──
    print(f"\n{'=' * 70}")
    print(f"  GENERATING EXCEL REPORT")
    print(f"{'=' * 70}")

    audit_meta = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "base_url": args.base_url,
        "locale": ", ".join(l for l, _ in locales),
        "max_pages": args.max_pages,
    }

    write_excel_report(all_issues, args.output, all_visited, audit_meta)

    # Optionally save JSON too
    if args.json_out:
        with open(args.json_out, "w", encoding="utf-8") as f:
            json.dump({
                "meta": audit_meta,
                "pages_visited": len(all_visited),
                "total_issues": len(all_issues),
                "issues": all_issues,
            }, f, ensure_ascii=False, indent=2)
        print(f"JSON results saved to: {args.json_out}")

    # Final summary
    print(f"\n{'=' * 70}")
    print(f"  AUDIT COMPLETE")
    print(f"{'=' * 70}")
    print(f"  Pages crawled: {len(all_visited)}")
    print(f"  Total issues:  {len(all_issues)}")

    by_cat = defaultdict(int)
    for i in all_issues:
        by_cat[i["category"]] += 1
    print(f"\n  By category:")
    for cat, count in sorted(by_cat.items(), key=lambda x: -x[1]):
        print(f"    {cat:25s} {count}")

    by_sev = defaultdict(int)
    for i in all_issues:
        by_sev[i["severity"]] += 1
    print(f"\n  By severity:")
    for sev in ["critical", "high", "medium", "low"]:
        print(f"    {sev:10s} {by_sev[sev]}")

    print(f"\n  Report: {args.output}")


if __name__ == "__main__":
    main()
